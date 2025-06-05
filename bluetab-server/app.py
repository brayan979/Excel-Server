from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
import tempfile
from werkzeug.utils import secure_filename


app = Flask(__name__)

# Configuración de CORS más permisiva
CORS(app, resources={
    r"/*": {
        "origins": os.getenv('CORS_ORIGINS', '*').split(','),
        "methods": os.getenv('CORS_METHODS', 'GET,POST,OPTIONS').split(','),
        "allow_headers": os.getenv('CORS_HEADERS', 'Content-Type,Authorization').split(',')
    }
})

# Configuración de archivos permitidos
ALLOWED_EXTENSIONS = set(os.getenv('ALLOWED_EXTENSIONS', 'xlsx').split(','))
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')
TEMPLATE_FILE = os.getenv('TEMPLATE_FILE', 'Plantilla.xlsx')

# Crear carpeta de uploads si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def procesar_datos_excel(archivo_datos, fecha_inicio, fecha_fin):
    """
    Procesa los datos del Excel según las fechas especificadas usando la plantilla existente
    """
    try:
        # Verificar que la plantilla existe
        if not os.path.exists(TEMPLATE_FILE):
            raise Exception("No se encuentra el archivo de plantilla")

        # Cargar la plantilla
        workbook = load_workbook(TEMPLATE_FILE)
        hoja = workbook['Hoja1']
        
        # Leer el archivo de datos
        df = pd.read_excel(archivo_datos)
        
        # Limpiar los datos
        df = df.drop(['HORAS','FECHAREGISTRO2', 'NºALBARAN','NºCLIENTE', 'NºREGISTRO'], axis=1)
        df.dropna(subset=['FECHAREGISTRO'], axis=0, inplace=True)
        df['FECHAREGISTRO'] = pd.to_datetime(df['FECHAREGISTRO'], format='%Y%m%d')
        df.IMPORTE = pd.to_numeric(df.IMPORTE, errors='coerce')
        df['HORAMOVIMIENTO'] = pd.to_datetime(df['HORAMOVIMIENTO'], format='%H:%M').dt.time
        df = df[(df['CONCEPTO'] != 'Apertura caja turno 1')]
        
        # Convertir fechas de string a datetime
        fecha_inicio_dt = pd.to_datetime(fecha_inicio)
        fecha_fin_dt = pd.to_datetime(fecha_fin)
        
        # Filtrar por rango de fechas
        df_filtrado = df[
            (df['FECHAREGISTRO'] >= fecha_inicio_dt) & 
            (df['FECHAREGISTRO'] <= fecha_fin_dt) & 
            (df['CONCEPTO'] != 'SE ARREPIENTE EL CLIENTE')
        ].copy()
        
        # Agrupar por día
        suma_por_dia = df_filtrado.groupby('FECHAREGISTRO')['IMPORTE'].sum()
        suma_por_diadf = pd.DataFrame(suma_por_dia)
        
        # Crear DataFrame con días completos del rango
        fechas_completas = pd.date_range(start=fecha_inicio_dt, end=fecha_fin_dt)
        df_completo = pd.DataFrame(fechas_completas, columns=['FECHAREGISTRO'])
        df_completo['Dia'] = df_completo['FECHAREGISTRO'].dt.day
        df_completo = df_completo.set_index('FECHAREGISTRO')
        
        # Combinar con los datos reales
        new_df = pd.concat([suma_por_dia.index.to_series().dt.day, suma_por_diadf.copy()], axis=1)
        new_df.columns = ['Dia', 'Importe']
        new_df = new_df.reset_index()
        
        new_df_completo = df_completo.combine_first(new_df.set_index('FECHAREGISTRO'))
        new_df_completo = new_df_completo.sort_index()
        new_df_completo['Importe'] = new_df_completo['Importe'].fillna(0)
        
        # Obtener lista de importes
        importes = new_df_completo['Importe'].to_list()
        
        # Escribir en la plantilla
        for i in range(len(importes)):
            hoja.cell(row=11+i, column=2, value=importes[i])
        
        # Guardar en archivo temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name, len(importes), df_filtrado.shape[0]
        
    except Exception as e:
        raise Exception(f"Error procesando los datos: {str(e)}")

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        return procesar()
    return jsonify({
        "status": "success",
        "message": "API de procesamiento de Excel",
        "endpoints": {
            "/": "GET/POST - Información de la API / Procesar archivo",
            "/procesar": "POST - Procesar archivo Excel"
        }
    })

def procesar():
    print("=== Nueva petición recibida ===")
    print("Método:", request.method)
    print("URL:", request.url)
    print("Headers:", dict(request.headers))
    print("Form:", request.form)
    print("Files:", request.files)
    print("============================")
        
    try:
        # Verificar que se hayan subido los archivos necesarios
        if 'archivo_datos' not in request.files:
            return jsonify({"error": "No se ha seleccionado el archivo de datos"}), 400
        
        archivo_datos = request.files['archivo_datos']
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        
        print("Archivo recibido:", archivo_datos.filename)
        print("Fecha inicio:", fecha_inicio)
        print("Fecha fin:", fecha_fin)
        
        if archivo_datos.filename == '':
            return jsonify({"error": "No se ha seleccionado ningún archivo"}), 400
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Debe especificar las fechas de inicio y fin"}), 400
        
        # Validar fechas
        try:
            datetime.strptime(fecha_inicio, '%Y-%m-%d')
            datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD"}), 400
        
        if archivo_datos and allowed_file(archivo_datos.filename):
            # Guardar archivo temporalmente
            filename = secure_filename(archivo_datos.filename)
            temp_path = os.path.join(UPLOAD_FOLDER, filename)
            archivo_datos.save(temp_path)
            
            try:
                # Procesar los datos
                archivo_resultado, num_dias, num_registros = procesar_datos_excel(
                    temp_path, fecha_inicio, fecha_fin
                )
                
                # Limpiar archivo temporal
                os.remove(temp_path)
                
                # Enviar archivo resultado
                return send_file(
                    archivo_resultado,
                    as_attachment=True,
                    download_name=f'reporte_{fecha_inicio}_{fecha_fin}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as e:
                # Limpiar archivo temporal en caso de error
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return jsonify({"error": str(e)}), 500
        
        else:
            return jsonify({"error": "Tipo de archivo no permitido. Solo se permiten archivos .xlsx"}), 400
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    print("=== Iniciando servidor ===")
    print("Rutas registradas:")
    for rule in app.url_map.iter_rules():
        print(f"{rule.endpoint}: {rule.methods} {rule}")
    print("========================")
    
    # Obtener configuración del servidor desde variables de entorno

    
    app.run(debug=True, host='0.0.0.0', port=3005)